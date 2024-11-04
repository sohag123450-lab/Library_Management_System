from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import PhotoImage
from tkinter import messagebox
from datetime import datetime, timedelta
import openpyxl
import pyxll
from openpyxl import load_workbook
import ast


def addbook():
    top.withdraw()
    new = Tk()
    file_path = r"C:\Users\Sabbir Ahmed\Desktop\final project\pythonProject98\for addbook.xlsx"
    A = openpyxl.load_workbook(file_path)
    B = A["addbook"]
    new.iconbitmap(r'books.ico')
    new.title('Library Management System')
    new.geometry('500x500+300+200')
    new.configure(bg='white')
    new.resizable(False, False)

    def submit():
        Book_id = bookid.get()
        Title = title.get()
        Author = author.get()
        Edition = edition.get()
        Price = price.get()
        if bookid and title and author and edition and price:
            B.append([Book_id, Title, Author, Edition, Price])
            A.save(file_path)
            messagebox.showinfo('Status', 'Data Submitted')
        else:
            messagebox.showerror('wrong', 'Field left Blanks')

    label4 = Label(new, text='Add Book', font=('Areal', 15, 'bold'), width=40, fg='white', bg='red')
    label4.place(x=8, y=40)
    label5 = Label(new, text='Book Id:', width=7, fg='white', bg='blue', font=('Areal', 15, 'bold'))
    label5.place(x=5, y=90)
    label6 = Label(new, text='Title:', width=7, fg='white', bg='blue', font=('Areal', 15, 'bold'))
    label6.place(x=5, y=150)
    label7 = Label(new, text='Author:', width=7, fg='white', bg='blue', font=('Areal', 15, 'bold'))
    label7.place(x=5, y=210)
    label8 = Label(new, text='Edition:', width=7, fg='white', bg='blue', font=('Areal', 15, 'bold'))
    label8.place(x=5, y=270)
    label9 = Label(new, text='Price:', width=7, fg='white', bg='blue', font=('Areal', 15, 'bold'))
    label9.place(x=5, y=330)
    bookid = Entry(new, width=25, bg='black', fg='white', bd='5', font=('Areal', 12, 'bold'))
    bookid.place(x=150, y=90)
    title = Entry(new, width=25, bg='black', fg='white', bd='5', font=('Areal', 12, 'bold'))
    title.place(x=150, y=150)
    author = Entry(new, width=25, bg='black', fg='white', bd='5', font=('Areal', 12, 'bold'))
    author.place(x=150, y=210)
    edition = Entry(new, width=25, bg='black', fg='white', bd='5', font=('Areal', 12, 'bold'))
    edition.place(x=150, y=270)
    price = Entry(new, width=25, bg='black', fg='white', bd='5', font=('Areal', 12, 'bold'))
    price.place(x=150, y=330)

    bottom = Button(new, text='Submit', bg='blue', fg='black', bd='5', command=submit)
    bottom.place(x=350, y=380)

    def Back():
        new.destroy()
        top.deiconify()

    button10 = Button(new, text='Back', bg='white', bd='5', command=Back)
    button10.place(x=350, y=420)


def addstudent():
    top.withdraw()
    new = Tk()
    file_path = r"C:\Users\Sabbir Ahmed\Desktop\final project\pythonProject98\for addstudent.xlsx"
    C = openpyxl.load_workbook(file_path)
    D = C["Addstudent"]
    new.iconbitmap(r'books.ico')
    new.title('Library Management System')
    new.geometry('500x500+300+200')
    new.configure(bg='white')
    new.resizable(False, False)

    def submit():
        Student_Id = stuid.get()
        Student_Name = stuname.get()
        Contact = contact.get()
        Batch = batch.get()
        Department = dept.get()
        University = uv.get()
        if stuid and stuname and contact and batch and dept and uv:
            D.append([Student_Id, Student_Name, Contact, Batch, Department, University])
            C.save(file_path)
            messagebox.showinfo('Status', 'Data Submitted')
        else:
            messagebox.showerror('wrong', 'Field left Blanks')

    label10 = Label(new, text='Add Student', width=40, bg='red', fg='white', font=('Areal', 15, 'bold'))
    label10.place(x=8, y=40)
    label11 = Label(new, text='Student Id:', width=11, fg='white', bg='blue', font=('Areal', 15, 'bold'))
    label11.place(x=5, y=90)
    label12 = Label(new, text='Student Name:', width=11, fg='white', bg='blue', font=('Areal', 15, 'bold'))
    label12.place(x=5, y=150)
    label13 = Label(new, text='Contact:', width=11, fg='white', bg='blue', font=('Areal', 15, 'bold'))
    label13.place(x=5, y=210)
    label14 = Label(new, text='Batch:', width=11, fg='white', bg='blue', font=('Areal', 15, 'bold'))
    label14.place(x=5, y=270)
    label15 = Label(new, text='Department:', width=11, fg='white', bg='blue', font=('Areal', 15, 'bold'))
    label15.place(x=5, y=330)
    label16 = Label(new, text='University:', width=11, fg='white', bg='blue', font=('Areal', 15, 'bold'))
    label16.place(x=5, y=380)
    stuid = Entry(new, width=25, fg='white', bg='black', bd='5', font=('Areal', 12, 'bold'))
    stuid.place(x=150, y=90)
    stuname = Entry(new, width=25, fg='white', bg='black', bd='5', font=('Areal', 12, 'bold'))
    stuname.place(x=150, y=150)
    contact = Entry(new, width=25, fg='white', bg='black', bd='5', font=('Areal', 12, 'bold'))
    contact.place(x=150, y=210)
    batch = Entry(new, width=25, fg='white', bg='black', bd='5', font=('Areal', 12, 'bold'))
    batch.place(x=150, y=270)
    dept = Entry(new, width=25, fg='white', bg='black', bd='5', font=('Areal', 12, 'bold'))
    dept.place(x=150, y=330)
    uv = Entry(new, width=25, fg='white', bg='black', bd='5', font=('Areal', 12, 'bold'))
    uv.place(x=150, y=380)
    bottom = Button(new, text='Submit', bg='blue', fg='black', bd='5', command=submit)
    bottom.place(x=350, y=440)

    def Back():
        new.destroy()
        top.deiconify()

    button11 = Button(new, text='Back', bg='white', bd='5', command=Back)
    button11.place(x=290, y=440)


top = Tk()
top.iconbitmap(r'books.ico')
top.title('Library Management System')
top.geometry('925x500+300+200')
top.configure(bg='black')
top.resizable(False, False)
img2 = PhotoImage(file='bgagain.png')
label2 = Label(top, image=img2, bg='white')
label2.place(x=10, y=0)
label3 = Label(text='Dashboard', fg='black', bg='blue', width=10, font=('areal', 20, 'bold', 'underline'))
label3.place(x=390, y=100)
addbook = Button(text='Add Book', fg='black', bg='gray', width=10, font=('areal', 20, 'bold', 'underline'),
                 command=addbook)
addbook.place(x=100, y=300)
addstudent = Button(text='Add Student', fg='black', bg='gray', width=10, font=('areal', 20, 'bold', 'underline'),
                    command=addstudent)
addstudent.place(x=390, y=300)


def bookreturn():
    def load_data():
        path = r"C:\Users\Sabbir Ahmed\Downloads\final project\final project\pythonProject98\for issue and return.xlsx"
        workbook = openpyxl.load_workbook(path)
        issue = workbook.active
        list_value = list(issue.values)

        for item in treeview.get_children():
            treeview.delete(item)

        treeview["columns"] = list_value[0]
        for col_name in list_value[0]:
            treeview.heading(col_name, text=col_name)

        for value_tuple in list_value[1:]:
            tags = ('returned',) if value_tuple[-1] == 'returned' else ()
            treeview.insert('', tk.END, values=value_tuple, tags=tags)

    def mark_as_returned():
        book_id = book_id_entry.get()
        if not book_id:
            messagebox.showwarning('Warning', 'Please enter a Book ID.')
            return

        path = r"C:\Users\Sabbir Ahmed\Desktop\final project\pythonProject98\for issue and return.xlsx"
        wb = load_workbook(path)
        ws = wb.active

        found = False
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == book_id:
                ws.cell(row=row[0].row, column=len(ws[1]), value='returned')
                found = True
                break

        if found:
            wb.save(path)
            wb.close()
            load_data()
        else:
            wb.close()
            messagebox.showwarning('Warning', 'Book ID not found')

    def delete_row():
        book_id = book_id_entry.get()
        if not book_id:
            messagebox.showwarning('Warning', 'Please enter a Book ID.')
            return

        path = r"C:\Users\Sabbir Ahmed\Desktop\final project\pythonProject98\for issue and return.xlsx"
        wb = load_workbook(path)
        ws = wb.active

        found = False
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            if row[0].value == book_id:
                ws.delete_rows(row[0].row, 1)
                found = True
                break

        if found:
            wb.save(path)
            wb.close()
            messagebox.showinfo('Success', 'Row deleted successfully')
            load_data()
        else:
            wb.close()
            messagebox.showwarning('Warning', 'Book ID not found')

    def search_book():
        search_id = book_id_entry.get()
        for item in treeview.get_children():
            treeview.delete(item)

        path = r"C:\Users\Sabbir Ahmed\Desktop\final project\pythonProject98\for issue and return.xlsx"
        workbook = openpyxl.load_workbook(path)
        issue = workbook.active

        list_value = list(issue.values)
        for value_tuple in list_value[1:]:
            if str(value_tuple[0]) == search_id:
                tags = ('returned',) if value_tuple[-1] == 'returned' else ()
                treeview.insert('', tk.END, values=value_tuple, tags=tags)
                return

    def on_treeview_click(event):
        selected_item = treeview.selection()
        if selected_item:
            item = treeview.item(selected_item)
            book_id_entry.delete(0, tk.END)
            book_id_entry.insert(0, item['values'][0])

    window = tk.Tk()
    window.iconbitmap(r"books.ico")
    window.title("Library Management System")
    window.geometry('1000x500')
    window.resizable(False, False)

    frame = ttk.Frame(window)
    frame.pack()

    book_idlabel = tk.Label(window, text="Enter Book ID:")
    book_idlabel.pack()

    book_id_entry = tk.Entry(window)
    book_id_entry.pack()

    search_button = tk.Button(window, text="Search", command=search_book)
    search_button.pack()

    delete_button = tk.Button(window, text="Delete Row", command=delete_row)
    delete_button.pack()

    ok_button = tk.Button(window, text="Return Book", command=mark_as_returned)
    ok_button.pack()
    def back():
        window.destroy()

    ok_button = tk.Button(window, text="Back", command=back)
    ok_button.pack()

    treeFrame = ttk.Frame(frame)
    treeFrame.grid(row=0, column=1, pady=10)
    treeScroll = ttk.Scrollbar(treeFrame)
    treeScroll.pack(side='right', fill='y')

    cols = ("Book_Id", "Student_Id", "Student_Name", "Issue_Date", "Return_Date", "Status")
    treeview = ttk.Treeview(treeFrame, show="headings", yscrollcommand=treeScroll.set, columns=cols, height=13)
    treeview.column('Book_Id', width=100)
    treeview.column('Student_Id', width=100)
    treeview.column('Student_Name', width=100)
    treeview.column('Issue_Date', width=100)
    treeview.column('Return_Date', width=100)
    treeview.column('Status', width=100)
    treeview.pack()
    treeScroll.configure(command=treeview.yview)

    treeview.tag_configure('returned', background='green')

    treeview.bind('<ButtonRelease-1>', on_treeview_click)

    load_data()

    window.mainloop()


bookreturn_button = tk.Button(top, text='Book Return', fg='black', bg='gray', width=10,
                              font=('Arial', 20, 'bold', 'underline'),
                              command=bookreturn)
bookreturn_button.place(x=240, y=200)




def issue():
    top.destroy()
    import issuebook1



issuebook = Button(text='Issue Book', fg='black', bg='gray', width=10, font=('areal', 20, 'bold', 'underline'),
                   command=issue)
issuebook.place(x=540, y=200)


#def show():
 #   def load_data():
  #      path = r"C:\Users\Sabbir Ahmed\Desktop\final project\pythonProject98\for addbook.xlsx"
   #     workbook = openpyxl.load_workbook(path)
    #    addbook = workbook.active

#        list_values = list(addbook.values)
 #       cols = list_values[0]
  #      tree = ttk.Treeview(window, column=cols, show="headings")
   #     for col_name in cols:
    #        tree.heading(col_name, text=col_name)
     #   tree.pack(expand=True, fill='y')

      #  for value_tuple in list_values[1:]:
       #     tree.insert('', END, values=value_tuple)

#    window = Tk()
 #   window.iconbitmap(r'books.ico')
  # window.title("Library Management System")
   # window.geometry('925x500+300+200')
   # window.resizable('False','False')

    #def back():
     #   window.destroy()

    #load_data()

   # button10 = Button(window, text='Back', width=5, fg='black', bg='red', command=back)
    #button10.place(x=0, y=0)

    #window.mainloop()


#showbook = Button(text='Show Book', fg='black', bg='gray', width=10, font=('areal', 20, 'bold', 'underline'),
 #                 command=show)
#showbook.place(x=700, y=200)

def exit():
    top.destroy()
exit= Button(text='Exit', fg='black', bg='gray', width=10, font=('areal', 20, 'bold', 'underline'), command=exit)
exit.place(x=700, y=300)

top.mainloop()