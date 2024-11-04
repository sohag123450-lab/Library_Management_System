import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
from openpyxl import load_workbook
from tkinter import filedialog
import openpyxl
import pyxll
from openpyxl import Workbook

# Create a new workbook and sheets if not exist
issue_file_path = r"C:\Users\Sabbir Ahmed\Downloads\final project\final project\pythonProject98\for issue and return.xlsx"
books_file_path = r"C:\Users\Sabbir Ahmed\Downloads\final project\final project\pythonProject98\for addbook.xlsx"

try:
    wb = openpyxl.load_workbook(issue_file_path)
    issue = wb.active
except FileNotFoundError:
    wb = Workbook()
    issue = wb.active
    issue.append(["BookID", "StudentID", "StudentName", "IssueDate", "ReturnDate"])
    wb.save(issue_file_path)

try:
    books_wb = openpyxl.load_workbook(books_file_path)
    addbook = books_wb.active
except FileNotFoundError:
    books_wb = Workbook()
    addbook = books_wb.active
    addbook.append(["BookID", "Title", "Author", "Edition", "Price"])
    books_wb.save(books_file_path)

def update_return_date(*args):
    issue_date_str = issue_date_entry.get()
    try:
        issue_date_obj = datetime.strptime(issue_date_str, "%Y-%m-%d")
        return_date_obj = issue_date_obj + timedelta(days=7)
        return_date_entry.config(state='normal')
        return_date_entry.delete(0, tk.END)
        return_date_entry.insert(0, return_date_obj.strftime("%Y-%m-%d"))
        return_date_entry.config(state='readonly')
    except ValueError:
        return_date_entry.config(state='normal')
        return_date_entry.delete(0, tk.END)
        return_date_entry.config(state='readonly')

def issue_book():
    book_id = book_id_entry.get()
    student_id = student_id_entry.get()
    student_name = student_name_entry.get()
    issue_date = issue_date_entry.get()
    return_date = return_date_entry.get()

    if not book_id or not student_id or not student_name or not issue_date or not return_date:
        messagebox.showerror("Error", "All fields are required!")
        return

    try:
        issue_date_obj = datetime.strptime(issue_date, "%Y-%m-%d")
        return_date_obj = datetime.strptime(return_date, "%Y-%m-%d")
    except ValueError:
        messagebox.showerror("Error", "Date format should be YYYY-MM-DD!")
        return

    if return_date_obj < issue_date_obj + timedelta(days=7):
        messagebox.showerror("Error", "Return date must be at least 7 days after the issue date!")
        return

    issue.append([book_id, student_id, student_name, issue_date, return_date])
    try:
        wb.save(issue_file_path)
        messagebox.showinfo("Success", "Book issued successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"Failed to save book issue details: {e}")
    wb.close()

def search_title():
    search_term = search_entry.get()
    for row in tree.get_children():
        tree.delete(row)
    for row in addbook.iter_rows(min_row=2, values_only=True):
        if search_term.lower() in row[1].lower():
            tree.insert("", tk.END, values=row)


root = tk.Tk()
root.title("Library Management System")
root.iconbitmap(r'books.ico')
root.title('Library Management System')
root.geometry('1000x700+300+200')
root.resizable(False, False)

# Issue Book Frame
frame = tk.Frame(root)
frame.pack(pady=20)

tk.Label(frame, text="Book ID:").grid(row=0, column=0)
book_id_entry = tk.Entry(frame)
book_id_entry.grid(row=0, column=1)

tk.Label(frame, text="Student ID:").grid(row=1, column=0)
student_id_entry = tk.Entry(frame)
student_id_entry.grid(row=1, column=1)

tk.Label(frame, text="Student Name:").grid(row=2, column=0)
student_name_entry = tk.Entry(frame)
student_name_entry.grid(row=2, column=1)

tk.Label(frame, text="Issue Date (YYYY-MM-DD):").grid(row=3, column=0)
issue_date_entry = tk.Entry(frame)
issue_date_entry.grid(row=3, column=1)
issue_date_entry.bind("<KeyRelease>", update_return_date)

tk.Label(frame, text="Return Date (YYYY-MM-DD):").grid(row=4, column=0)
return_date_entry = tk.Entry(frame)
return_date_entry.grid(row=4, column=1)
return_date_entry.config(state='readonly')

tk.Button(frame, text="Issue Book", command=issue_book).grid(row=5, columnspan=2, pady=10)
def back():
    root.destroy()
    import dashboard
tk.Button(frame, text="Back",command=back).grid(row=6,columnspan=3,pady=10)

# Treeview for Books Data
tree_frame = tk.Frame(root)
tree_frame.pack(pady=20)

tree_scroll = tk.Scrollbar(tree_frame)
tree_scroll.pack(side=tk.RIGHT, fill=tk.Y)

tree = ttk.Treeview(tree_frame, yscrollcommand=tree_scroll.set, columns=("BookID", "Title", "Author", "Edition", "Price"), show='headings')
tree.pack()

tree_scroll.config(command=tree.yview)

tree.heading("BookID", text="BookID")
tree.heading("Title", text="Title")
tree.heading("Author", text="Author")
tree.heading("Edition", text="Edition")
tree.heading("Price", text="Price")

# Populate Treeview
for row in addbook.iter_rows(min_row=2, values_only=True):
    tree.insert("", tk.END, values=row)

# Search Frame
search_frame = tk.Frame(root)
search_frame.pack(pady=20)

tk.Label(search_frame, text="Search Title:").pack(side=tk.LEFT)
search_entry = tk.Entry(search_frame)
search_entry.pack(side=tk.LEFT, padx=10)
tk.Button(search_frame, text="Search", command=search_title).pack(side=tk.LEFT)

root.mainloop()
